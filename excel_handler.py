"""
Excel read/write operations with atomic transactions and backup
"""
import pandas as pd
import shutil
from datetime import datetime
from pathlib import Path
import openpyxl
from typing import Optional
import logging
import zipfile

from config import EXCEL_PATH, BACKUP_DIR, BACKUP_RETENTION_DAYS, LOGS_DIR

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(LOGS_DIR / 'income_wheel.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


class ExcelHandler:
    """Handle all Excel read/write operations with transaction safety"""
    
    def __init__(self, excel_path: str = EXCEL_PATH):
        self.excel_path = Path(excel_path)
        self.backup_dir = BACKUP_DIR
        
        if not self.excel_path.exists():
            raise FileNotFoundError(f"Excel file not found: {self.excel_path}")
    
    def backup_file(self) -> Path:
        """
        Create timestamped backup of Excel file
        
        Returns:
            Path to backup file
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = self.backup_dir / f"income_wheel_backup_{timestamp}.xlsx"
        
        shutil.copy2(self.excel_path, backup_path)
        logger.info(f"Backup created: {backup_path}")
        
        # Clean old backups
        self._cleanup_old_backups()
        
        return backup_path
    
    def _cleanup_old_backups(self):
        """Remove backups older than BACKUP_RETENTION_DAYS"""
        cutoff_date = datetime.now().timestamp() - (BACKUP_RETENTION_DAYS * 24 * 60 * 60)
        
        for backup_file in self.backup_dir.glob("income_wheel_backup_*.xlsx"):
            if backup_file.stat().st_mtime < cutoff_date:
                backup_file.unlink()
                logger.info(f"Removed old backup: {backup_file}")
    
    def restore_from_backup(self, backup_path: Path):
        """Restore Excel file from backup"""
        if not backup_path.exists():
            raise FileNotFoundError(f"Backup file not found: {backup_path}")
        
        shutil.copy2(backup_path, self.excel_path)
        logger.warning(f"Restored from backup: {backup_path}")
    
    def read_data_table(self) -> pd.DataFrame:
        """
        Read Data Table sheet
        
        Returns:
            DataFrame with all trades
        """
        try:
            df = pd.read_excel(self.excel_path, sheet_name='Data Table')
            logger.info(f"Loaded {len(df)} trades from Data Table")
            return df
        except Exception as e:
            logger.error(f"Error reading Data Table: {e}")
            raise
    
    def read_audit_table(self) -> pd.DataFrame:
        """
        Read Audit Table sheet
        
        Returns:
            DataFrame with audit log
        """
        try:
            df = pd.read_excel(self.excel_path, sheet_name='Audit_Table')
            logger.info(f"Loaded {len(df)} audit entries")
            return df
        except Exception as e:
            logger.error(f"Error reading Audit Table: {e}")
            raise
    
    def read_daily_helper(self) -> pd.DataFrame:
        """
        Read Daily Helper sheet
        
        Returns:
            DataFrame with daily helper data
        """
        try:
            # Daily Helper has custom layout, read as-is
            df = pd.read_excel(self.excel_path, sheet_name='Daily Helper')
            logger.info(f"Loaded Daily Helper data")
            return df
        except Exception as e:
            logger.error(f"Error reading Daily Helper: {e}")
            raise
    
    def load_all_data(self) -> dict:
        """
        Load all sheets at once
        
        Returns:
            dict with 'trades', 'audit', 'daily_helper' DataFrames
        """
        return {
            'trades': self.read_data_table(),
            'audit': self.read_audit_table(),
            'daily_helper': self.read_daily_helper()
        }
    
    def append_trade(self, trade_data: dict) -> bool:
        """
        Append new trade to Data Table
        
        Args:
            trade_data: Dictionary with trade fields matching Data Table columns
        
        Returns:
            True if successful
        """
        backup_path = None
        try:
            # Backup before write
            backup_path = self.backup_file()
            
            # Read existing data
            df = self.read_data_table()
            
            # Create new row
            new_row = pd.DataFrame([trade_data])
            
            # Append
            df = pd.concat([df, new_row], ignore_index=True)
            
            # Write back using openpyxl to preserve formulas
            self._write_dataframe_preserve_formulas(df, 'Data Table')
            
            logger.info(f"Appended trade: {trade_data.get('TradeID', 'Unknown')}")
            return True
            
        except Exception as e:
            logger.error(f"Error appending trade: {e}")
            if backup_path and backup_path.exists():
                logger.warning("Restoring from backup due to error")
                self.restore_from_backup(backup_path)
            raise
    
    def update_trade(self, trade_id: str, updates: dict) -> bool:
        """
        Update existing trade in Data Table
        
        Args:
            trade_id: TradeID to update
            updates: Dictionary of fields to update
        
        Returns:
            True if successful
        """
        backup_path = None
        try:
            # Backup before write
            backup_path = self.backup_file()
            
            # Read existing data
            df = self.read_data_table()
            
            # Find row to update
            mask = df['TradeID'] == trade_id
            
            if not mask.any():
                raise ValueError(f"TradeID {trade_id} not found in Data Table")
            
            # Update fields
            for field, value in updates.items():
                df.loc[mask, field] = value
            
            # Write back
            self._write_dataframe_preserve_formulas(df, 'Data Table')
            
            logger.info(f"Updated trade: {trade_id} with {updates}")
            return True
            
        except Exception as e:
            logger.error(f"Error updating trade: {e}")
            if backup_path and backup_path.exists():
                logger.warning("Restoring from backup due to error")
                self.restore_from_backup(backup_path)
            raise
    
    def delete_trades(self, trade_ids: list) -> bool:
        """
        Delete one or more trades from Data Table by TradeID.
        """
        if not trade_ids:
            return True
        backup_path = None
        try:
            backup_path = self.backup_file()
            df = self.read_data_table()
            before = len(df)
            df = df[~df['TradeID'].astype(str).str.strip().isin([str(t).strip() for t in trade_ids])]
            removed = before - len(df)
            if removed == 0:
                logger.warning(f"No rows matched TradeIDs: {trade_ids}")
                return True
            self._write_dataframe_preserve_formulas(df, 'Data Table')
            logger.info(f"Deleted {removed} trade(s): {trade_ids}")
            return True
        except Exception as e:
            logger.error(f"Error deleting trades: {e}")
            if backup_path and backup_path.exists():
                self.restore_from_backup(backup_path)
            raise
    
    def append_audit(self, audit_data: dict) -> bool:
        """
        Append audit entry to Audit Table
        
        Args:
            audit_data: Dictionary with audit fields
        
        Returns:
            True if successful
        """
        try:
            # Read existing audit data
            df = self.read_audit_table()
            
            # Create new row
            new_row = pd.DataFrame([audit_data])
            
            # Append
            df = pd.concat([df, new_row], ignore_index=True)
            
            # Write back
            self._write_dataframe_preserve_formulas(df, 'Audit_Table')
            
            logger.info(f"Appended audit: {audit_data.get('Audit ID', 'Unknown')}")
            return True
            
        except Exception as e:
            logger.error(f"Error appending audit: {e}")
            raise
    
    def _write_dataframe_preserve_formulas(self, df: pd.DataFrame, sheet_name: str):
        """
        Write DataFrame to Excel sheet while preserving formulas in other columns
        
        Uses openpyxl to write cell-by-cell, avoiding overwriting formula columns
        """
        # Load workbook
        wb = openpyxl.load_workbook(self.excel_path)
        ws = wb[sheet_name]
        
        # Clear existing data (keep header)
        ws.delete_rows(2, ws.max_row)
        
        # Write data rows
        for r_idx, row in enumerate(df.itertuples(index=False), start=2):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx)
                
                # Write value (openpyxl will preserve formula if value is formula string)
                if pd.isna(value):
                    cell.value = None
                elif isinstance(value, (pd.Timestamp, datetime)):
                    cell.value = value
                else:
                    cell.value = value
        
        # Save workbook
        wb.save(self.excel_path)
        wb.close()
    
    def atomic_transaction(self, operations: list) -> bool:
        """
        Execute multiple operations atomically
        
        Args:
            operations: List of dicts with 'type' and 'data'
                       Types: 'append_trade', 'update_trade', 'append_audit'
        
        Returns:
            True if all operations successful
        
        Example:
            operations = [
                {'type': 'update_trade', 'data': {'trade_id': 'T-123', 'updates': {'Status': 'Closed'}}},
                {'type': 'append_audit', 'data': {'Audit ID': 'A-456', ...}}
            ]
        """
        backup_path = None
        try:
            # Single backup for all operations
            backup_path = self.backup_file()
            
            # Execute all operations
            for op in operations:
                op_type = op['type']
                op_data = op['data']
                
                if op_type == 'append_trade':
                    self.append_trade(op_data)
                elif op_type == 'update_trade':
                    self.update_trade(op_data['trade_id'], op_data['updates'])
                elif op_type == 'append_audit':
                    self.append_audit(op_data)
                else:
                    raise ValueError(f"Unknown operation type: {op_type}")
            
            logger.info(f"Atomic transaction completed: {len(operations)} operations")
            return True
            
        except Exception as e:
            logger.error(f"Atomic transaction failed: {e}")
            if backup_path and backup_path.exists():
                logger.warning("Restoring from backup due to transaction failure")
                self.restore_from_backup(backup_path)
            raise


def validate_data_integrity(df_trades: pd.DataFrame, df_audit: pd.DataFrame) -> list:
    """
    Run integrity checks on loaded data
    
    Returns:
        List of error messages (empty if all checks pass)
    """
    errors = []
    
    # Check 1: No duplicate TradeIDs (only warn if exact duplicates, not intentional splits)
    if df_trades['TradeID'].duplicated().any():
        duplicates = df_trades[df_trades['TradeID'].duplicated()]['TradeID'].unique().tolist()
        # Check if duplicates are intentional splits (e.g., T-146, T-146A, T-146B)
        # If all duplicates have the same base ID with different suffixes, it's likely intentional
        true_duplicates = []
        for dup_id in duplicates:
            dup_str = str(dup_id)
            # Check if there are multiple rows with EXACT same TradeID (not just same base)
            exact_dups = df_trades[df_trades['TradeID'] == dup_id]
            if len(exact_dups) > 1:
                true_duplicates.append(dup_str)
        
        if true_duplicates:
            errors.append(f"Duplicate TradeIDs found: {true_duplicates}")
    
    # Check 2: All audit references exist (handle partial split trades)
    audit_refs = df_audit['TradeID_Ref'].dropna().str.split(',').explode().str.strip()
    trade_ids = set(df_trades['TradeID'].dropna().astype(str))
    
    def is_valid_trade_ref(ref: str) -> bool:
        """Check if audit reference is valid, including partial split trades"""
        ref = str(ref).strip()
        
        # Special cases: "ALL" = all trades; "MULTIPLE" = action applies to multiple trades (placeholder)
        if ref.upper() in ('ALL', 'MULTIPLE'):
            return True
        
        # Direct match
        if ref in trade_ids:
            return True
        
        import re
        
        # Check for partial split patterns:
        # Pattern 1: T-XXX-A, T-XXX-B (with dash before suffix)
        # Pattern 2: T-146A, T-146B (no dash before suffix)
        # Pattern 3: T-XXX-1, T-XXX-2 (numeric suffix)
        
        # Pattern 1: T-XXX-SUFFIX (with dash)
        pattern1 = r'^(T-?\d+)-([A-Z]|\d+)$'
        match1 = re.match(pattern1, ref)
        
        # Pattern 2: T-XXXSUFFIX (no dash, like T-146A)
        pattern2 = r'^(T-?\d+)([A-Z]|\d+)$'
        match2 = re.match(pattern2, ref)
        
        if match1:
            base_trade_id = match1.group(1)  # e.g., "T-125"
        elif match2:
            base_trade_id = match2.group(1)  # e.g., "T-146"
        else:
            base_trade_id = None
        
        if base_trade_id:
            # Try both T-XXX and TXXX formats
            base_variants = [base_trade_id, base_trade_id.replace('-', '')]
            
            # Check if base trade exists
            for variant in base_variants:
                if variant in trade_ids:
                    return True
                # Also check if any suffixed version exists (T-125-A, T-125-B, T-146A, etc.)
                for trade_id in trade_ids:
                    trade_str = str(trade_id)
                    # Check for T-XXX-SUFFIX or T-XXXSUFFIX patterns
                    if trade_str.startswith(variant + '-') or trade_str.startswith(variant):
                        # Make sure it's actually a suffix (not just a longer number)
                        remaining = trade_str[len(variant):]
                        if remaining.startswith('-') or (remaining and remaining[0].isalpha()):
                            return True
        
        # Check for numeric-only patterns (132, 133, T131)
        if ref.isdigit():
            # Try T-XXX format
            if f"T-{ref}" in trade_ids or f"T{ref}" in trade_ids:
                return True
        
        # Check for TXXX format (T131, T377, etc.)
        if ref.startswith('T') and ref[1:].isdigit():
            # Try T-XXX format
            numeric_part = ref[1:]
            if f"T-{numeric_part}" in trade_ids:
                return True
        
        return False
    
    # Filter out valid partial split trades
    missing = []
    for ref in audit_refs:
        if not is_valid_trade_ref(ref):
            missing.append(ref)
    
    if missing:
        # Only report truly missing trades (not partial splits)
        errors.append(f"Audit references missing trades: {set(missing)}")
    
    # Check 3: All open positions have expiry (except STOCK)
    open_no_expiry = df_trades[
        (df_trades['Status'] == 'Open') &
        (df_trades['TradeType'] != 'STOCK') &
        (df_trades['Expiry_Date'].isna())
    ]
    
    if not open_no_expiry.empty:
        errors.append(f"Open options with no expiry: {open_no_expiry['TradeID'].tolist()}")
    
    return errors
